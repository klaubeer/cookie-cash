from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ─── Paleta ───────────────────────────────────────────────────────────────────
COR_HEADER     = "2D6A4F"
COR_RESUMO_BG  = "1B4332"
COR_BRANCO     = "FFFFFF"
COR_RECEITA_BG = "D8F3DC"
COR_DESPESA_BG = "FFCCD5"
FONTE          = "Arial"

def borda(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, value=None, bg=None, fg="000000", bold=False,
         size=10, align="center", fmt=None, italic=False, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONTE, bold=bold, color=fg, size=size, italic=italic)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = borda()
    if fmt:
        c.number_format = fmt
    return c

# ══════════════════════════════════════════════════════════════════════════════
# ABA 1 — LANÇAMENTOS
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Lançamentos"
ws1.sheet_view.showGridLines = False

headers = ["Data", "Tipo", "Descrição", "Valor (R$)", "Origem", "Timestamp"]
widths  = [14, 12, 40, 14, 10, 22]

for col, (h, w) in enumerate(zip(headers, widths), 1):
    cell(ws1, 1, col, h, bg=COR_HEADER, fg=COR_BRANCO, bold=True, size=11)
    ws1.column_dimensions[get_column_letter(col)].width = w
ws1.row_dimensions[1].height = 28

dados = [
    ("27/03/2026", "RECEITA",  "Joana — 4 cookies",                   60.00,  "texto"),
    ("27/03/2026", "RECEITA",  "Ana — 6 cookies",                     90.00,  "áudio"),
    ("27/03/2026", "DESPESA",  "Mercadão — farinha, ovos, manteiga",  87.50,  "foto"),
    ("26/03/2026", "RECEITA",  "Maria — caixa 12 cookies",           150.00,  "texto"),
    ("26/03/2026", "DESPESA",  "Atacadão — chocolate, açúcar",       112.30,  "foto"),
    ("25/03/2026", "RECEITA",  "Carla — 3 cookies",                   45.00,  "texto"),
    ("25/03/2026", "DESPESA",  "Mercado Bom Preço — leite condensado", 34.90, "foto"),
    ("24/03/2026", "RECEITA",  "Patrícia — encomenda festa",         280.00,  "áudio"),
    ("23/03/2026", "RECEITA",  "Julia — 5 cookies",                   75.00,  "texto"),
    ("23/03/2026", "DESPESA",  "Atacadão — embalagens",               48.00,  "foto"),
    ("10/02/2026", "RECEITA",  "Sandra — 8 cookies",                 120.00,  "texto"),
    ("05/02/2026", "DESPESA",  "Mercadão — insumos fevereiro",       210.00,  "foto"),
]

for i, (data, tipo, desc, valor, origem) in enumerate(dados, 2):
    ts = f"{data[6:]}-{data[3:5]}-{data[0:2]}T{'09:00:00' if i % 2 == 0 else '14:30:00'}"
    bg = COR_RECEITA_BG if tipo == "RECEITA" else COR_DESPESA_BG
    alt = "E8F8EC" if tipo == "RECEITA" else "FFE5EA"
    fill = bg if i % 2 == 0 else alt
    tipo_cor = "1B5E20" if tipo == "RECEITA" else "B71C1C"

    cell(ws1, i, 1, data,   bg=fill)
    cell(ws1, i, 2, tipo,   bg=fill, fg=tipo_cor, bold=True)
    cell(ws1, i, 3, desc,   bg=fill, align="left")
    cell(ws1, i, 4, valor,  bg=fill, fmt='#,##0.00')
    cell(ws1, i, 5, origem, bg=fill)
    cell(ws1, i, 6, ts,     bg=fill)
    ws1.row_dimensions[i].height = 22

# Total
tr = len(dados) + 2
cell(ws1, tr, 3, "SALDO DO PERÍODO", align="right", bold=True)
saldo = cell(ws1, tr, 4,
    f'=SUMIF(B2:B{tr-1},"RECEITA",D2:D{tr-1})-SUMIF(B2:B{tr-1},"DESPESA",D2:D{tr-1})',
    bg="B7E4C7", fg="1B5E20", bold=True, size=11, fmt='#,##0.00')
ws1.row_dimensions[tr].height = 26
ws1.freeze_panes = "A2"

# ══════════════════════════════════════════════════════════════════════════════
# ABA 2 — RESUMO
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Resumo")
ws2.sheet_view.showGridLines = False

for col, w in enumerate([22, 22, 22, 18], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w

# ── Título ──
ws2.merge_cells("A1:D1")
t = ws2.cell(row=1, column=1, value="🍪  Cookie Finance — Resumo")
t.font = Font(name=FONTE, bold=True, size=14, color=COR_BRANCO)
t.fill = PatternFill("solid", fgColor=COR_RESUMO_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 36

# ── Seletor de período ──
ws2.row_dimensions[2].height = 8  # espaço

ws2.merge_cells("A3:D3")
label_secao = ws2.cell(row=3, column=1, value="PERÍODO DE ANÁLISE")
label_secao.font = Font(name=FONTE, bold=True, size=9, color="888888")
label_secao.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[3].height = 18

# Label + dropdown do período
cell(ws2, 4, 1, "Período", bg="F0F0F0", bold=True, align="left")
periodo_cell = cell(ws2, 4, 2, "Mês atual", bg=COR_BRANCO, align="left")
ws2.row_dimensions[4].height = 26

dv = DataValidation(
    type="list",
    formula1='"Mês atual,Ano atual,Últimos 7 dias,Últimos 30 dias,Personalizado"',
    allow_blank=False,
    showDropDown=False,
)
dv.error = "Escolha uma opção da lista"
dv.errorTitle = "Valor inválido"
ws2.add_data_validation(dv)
dv.add("B4")

# Células De / Até (calculadas automaticamente)
# B5 = data início, C5 = data fim
cell(ws2, 5, 1, "De", bg="F0F0F0", bold=True)
de_cell = cell(ws2, 5, 2,
    '=IFS(B4="Mês atual",DATE(YEAR(TODAY()),MONTH(TODAY()),1),'
    'B4="Ano atual",DATE(YEAR(TODAY()),1,1),'
    'B4="Últimos 7 dias",TODAY()-7,'
    'B4="Últimos 30 dias",TODAY()-30,'
    'B4="Personalizado",B7)',
    bg=COR_BRANCO, fmt="DD/MM/YYYY")
ws2.row_dimensions[5].height = 26

cell(ws2, 6, 1, "Até", bg="F0F0F0", bold=True)
ate_cell = cell(ws2, 6, 2,
    '=IFS(B4="Mês atual",EOMONTH(TODAY(),0),'
    'B4="Ano atual",DATE(YEAR(TODAY()),12,31),'
    'B4="Últimos 7 dias",TODAY(),'
    'B4="Últimos 30 dias",TODAY(),'
    'B4="Personalizado",C7)',
    bg=COR_BRANCO, fmt="DD/MM/YYYY")
ws2.row_dimensions[6].height = 26

# Campos de data manual (só usados quando Personalizado)
cell(ws2, 7, 1, "↳ Personalizado:", bg="FFFDE7", bold=False, size=9, fg="888888", align="left")
cell(ws2, 7, 2, None, bg="FFFDE7", fmt="DD/MM/YYYY")   # usuária digita aqui
cell(ws2, 7, 3, None, bg="FFFDE7", fmt="DD/MM/YYYY")   # usuária digita aqui
nota_p = ws2.cell(row=7, column=2)
nota_p.value = None
ws2.cell(row=7, column=1).alignment = Alignment(horizontal="left", vertical="center")
# Pequena nota orientativa
ws2.merge_cells("D7:D7")
hint = ws2.cell(row=7, column=4, value='← De  |  Até →')
hint.font = Font(name=FONTE, size=8, color="AAAAAA", italic=True)
hint.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[7].height = 22

# ── Separador ──
ws2.row_dimensions[8].height = 10

# ── Resultados ──
ws2.merge_cells("A9:D9")
label_res = ws2.cell(row=9, column=1, value="RESULTADOS")
label_res.font = Font(name=FONTE, bold=True, size=9, color="888888")
label_res.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[9].height = 18

FORMULA_BASE = (
    '=SUMPRODUCT('
    '(Lançamentos!$A$2:$A$1000>=$B$5)*'
    '(Lançamentos!$A$2:$A$1000<=$B$6)*'
    '(Lançamentos!$B$2:$B$1000="{tipo}")*'
    'Lançamentos!$D$2:$D$1000)'
)

cell(ws2, 10, 1, "RECEITAS", bg=COR_RECEITA_BG, fg="1B5E20", bold=True, size=12)
cell(ws2, 10, 2, FORMULA_BASE.format(tipo="RECEITA"),
     bg=COR_RECEITA_BG, fg="1B5E20", bold=True, size=12, fmt='"R$" #,##0.00')
ws2.row_dimensions[10].height = 32

cell(ws2, 11, 1, "DESPESAS", bg=COR_DESPESA_BG, fg="B71C1C", bold=True, size=12)
cell(ws2, 11, 2, FORMULA_BASE.format(tipo="DESPESA"),
     bg=COR_DESPESA_BG, fg="B71C1C", bold=True, size=12, fmt='"R$" #,##0.00')
ws2.row_dimensions[11].height = 32

ws2.row_dimensions[12].height = 8

cell(ws2, 13, 1, "SALDO", bg=COR_HEADER, fg=COR_BRANCO, bold=True, size=13)
cell(ws2, 13, 2, "=B10-B11",
     bg=COR_HEADER, fg=COR_BRANCO, bold=True, size=13, fmt='"R$" #,##0.00')
ws2.row_dimensions[13].height = 36

# Nota
ws2.row_dimensions[15].height = 16
ws2.merge_cells("A15:D15")
nota = ws2.cell(row=15, column=1,
    value="* Atualizado automaticamente. Para período personalizado, preenha as células De/Até na linha 7.")
nota.font = Font(name=FONTE, size=8, color="AAAAAA", italic=True)
nota.alignment = Alignment(horizontal="left", vertical="center")

# ── Salvar ──
out = "C:/Users/cheff/cookie-cash/preview_planilha.xlsx"
wb.save(out)
print(f"Salvo: {out}")
